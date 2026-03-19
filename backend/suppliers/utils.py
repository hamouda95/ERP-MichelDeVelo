"""
Utilitaires pour le module Achat
"""
import csv
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.core.files.uploadedfile import InMemoryUploadedFile
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

class DocumentParser:
    """Parseur de documents pour les factures et bons de livraison"""
    
    @staticmethod
    def parse_excel(file):
        """Parser un fichier Excel"""
        try:
            workbook = load_workbook(filename=file, read_only=True)
            sheet = workbook.active
            
            items = []
            headers = {}
            
            # Trouver les en-têtes
            for row in sheet.iter_rows(max_row=2):
                if any(cell.value for cell in row):
                    for i, cell in enumerate(row):
                        if cell.value:
                            headers[str(cell.value).strip().lower()] = i
                    break
            
            # Parser les lignes de données
            for row in sheet.iter_rows(min_row=2):
                if not any(cell.value for cell in row):
                    continue
                    
                row_data = {}
                for header, col_idx in headers.items():
                    if col_idx < len(row) and row[col_idx].value:
                        row_data[header] = row[col_idx].value
                
                # Extraire les informations de produit
                if 'produit' in row_data or 'article' in row_data or 'désignation' in row_data:
                    item = {
                        'product_name': row_data.get('produit') or row_data.get('article') or row_data.get('désignation', ''),
                        'reference': row_data.get('référence') or row_data.get('ref') or '',
                        'quantity': DocumentParser._extract_quantity(row_data),
                        'unit_price': DocumentParser._extract_price(row_data),
                    }
                    items.append(item)
            
            workbook.close()
            return items
            
        except Exception as e:
            logger.error(f"Erreur parsing Excel: {e}")
            return []
    
    @staticmethod
    def parse_csv(file):
        """Parser un fichier CSV"""
        try:
            content = file.read().decode('utf-8-sig')
            csv_file = io.StringIO(content)
            reader = csv.DictReader(csv_file)
            
            items = []
            for row in reader:
                item = {
                    'product_name': row.get('produit') or row.get('article') or row.get('désignation', ''),
                    'reference': row.get('référence') or row.get('ref') or '',
                    'quantity': DocumentParser._extract_quantity(row),
                    'unit_price': DocumentParser._extract_price(row),
                }
                items.append(item)
            
            return items
            
        except Exception as e:
            logger.error(f"Erreur parsing CSV: {e}")
            return []
    
    @staticmethod
    def _extract_quantity(row_data):
        """Extraire la quantité depuis différentes colonnes possibles"""
        for key in ['quantité', 'quantite', 'qté', 'qte', 'quantity', 'qty']:
            if key in row_data and row_data[key]:
                try:
                    return Decimal(str(row_data[key]))
                except:
                    continue
        return Decimal('0')
    
    @staticmethod
    def _extract_price(row_data):
        """Extraire le prix unitaire depuis différentes colonnes possibles"""
        for key in ['prix unitaire', 'prix', 'pu', 'unit_price', 'price']:
            if key in row_data and row_data[key]:
                try:
                    # Nettoyer le prix (supprimer €, espaces, etc.)
                    price_str = str(row_data[key]).replace('€', '').replace(' ', '').strip()
                    return Decimal(price_str)
                except:
                    continue
        return Decimal('0')

class StockChecker:
    """Vérificateur de stocks et alertes"""
    
    @staticmethod
    def check_stock_levels(products, store=None):
        """Vérifier les niveaux de stock et générer des alertes"""
        alerts = []
        
        for product in products:
            if not product.is_active:
                continue
                
            # Récupérer le stock selon le magasin
            if store == 'ville_avray':
                current_stock = product.stock_ville_avray
            elif store == 'garches':
                current_stock = product.stock_garches
            else:
                current_stock = product.stock_ville_avray + product.stock_garches
            
            # Vérifier les alertes
            if current_stock <= product.alert_stock:
                alerts.append({
                    'product': product,
                    'current_stock': current_stock,
                    'alert_stock': product.alert_stock,
                    'severity': 'critical' if current_stock == 0 else 'warning',
                    'message': f"Stock critique: {product.name} - {current_stock} unités (alerte: {product.alert_stock})"
                })
        
        return alerts
    
    @staticmethod
    def generate_stock_report(products, store=None):
        """Générer un rapport de stock détaillé"""
        report = {
            'total_products': len([p for p in products if p.is_active]),
            'low_stock_products': 0,
            'out_of_stock_products': 0,
            'total_stock_value': 0,
            'alerts': [],
            'categories': {}
        }
        
        for product in products:
            if not product.is_active:
                continue
                
            # Récupérer le stock selon le magasin
            if store == 'ville_avray':
                current_stock = product.stock_ville_avray
            elif store == 'garches':
                current_stock = product.stock_garches
            else:
                current_stock = product.stock_ville_avray + product.stock_garches
            
            # Calculer la valeur du stock
            stock_value = current_stock * float(product.price_ht or 0)
            report['total_stock_value'] += stock_value
            
            # Catégoriser
            category = product.category.name if product.category else 'Non catégorisé'
            if category not in report['categories']:
                report['categories'][category] = {
                    'count': 0,
                    'stock_value': 0,
                    'low_stock': 0
                }
            
            report['categories'][category]['count'] += 1
            report['categories'][category]['stock_value'] += stock_value
            
            # Vérifier les alertes
            if current_stock <= 0:
                report['out_of_stock_products'] += 1
                report['categories'][category]['low_stock'] += 1
                report['alerts'].append({
                    'product': product,
                    'current_stock': current_stock,
                    'alert_stock': product.alert_stock,
                    'severity': 'critical',
                    'message': f"RUPTURE: {product.name} - {current_stock} unités"
                })
            elif current_stock <= product.alert_stock:
                report['low_stock_products'] += 1
                report['categories'][category]['low_stock'] += 1
                report['alerts'].append({
                    'product': product,
                    'current_stock': current_stock,
                    'alert_stock': product.alert_stock,
                    'severity': 'warning',
                    'message': f"STOCK FAIBLE: {product.name} - {current_stock} unités (alerte: {product.alert_stock})"
                })
        
        return report

class PurchaseOrderValidator:
    """Validateur de commandes d'achat"""
    
    @staticmethod
    def validate_order(order_data, products, suppliers):
        """Valider une commande d'achat"""
        errors = []
        warnings = []
        
        # Vérifier le fournisseur
        if not order_data.get('supplier'):
            errors.append("Le fournisseur est obligatoire")
        else:
            try:
                supplier = suppliers.get(id=order_data['supplier'])
                if not supplier.is_active:
                    errors.append("Le fournisseur sélectionné n'est pas actif")
            except:
                errors.append("Fournisseur invalide")
        
        # Vérifier les articles
        items = order_data.get('items', [])
        if not items:
            errors.append("La commande doit contenir au moins un article")
        
        for i, item in enumerate(items):
            item_errors = []
            
            # Vérifier le produit
            if not item.get('product'):
                item_errors.append("Le produit est obligatoire")
            else:
                try:
                    product = products.get(id=item['product'])
                    if not product.is_active:
                        item_errors.append("Le produit sélectionné n'est pas actif")
                except:
                    item_errors.append("Produit invalide")
            
            # Vérifier la quantité
            if not item.get('quantity') or item['quantity'] <= 0:
                item_errors.append("La quantité doit être supérieure à 0")
            
            # Vérifier le prix
            if not item.get('unit_price') or item['unit_price'] <= 0:
                item_errors.append("Le prix unitaire doit être supérieur à 0")
            
            if item_errors:
                errors.extend([f"Article {i+1}: {error}" for error in item_errors])
        
        # Vérifications additionnelles
        if order_data.get('expected_delivery_date'):
            from datetime import datetime, date
            try:
                delivery_date = datetime.strptime(order_data['expected_delivery_date'], '%Y-%m-%d').date()
                if delivery_date < date.today():
                    warnings.append("La date de livraison est antérieure à aujourd'hui")
            except:
                errors.append("Format de date de livraison invalide")
        
        return {
            'is_valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }
